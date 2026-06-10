#!/usr/bin/env python3
"""Assemble 150 character asset prompts using v2.0 template."""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============ v2.0 Fixed Templates ============

# Route A - Standard
STRUCTURE_A = "A 16:9 character asset sheet. Left 1/3 is a face close-up showing full head to collar. Right 2/3 has three full-body standing views side by side: front, side, back."

# Route B - Anti-CG (period drama)
STRUCTURE_B = "A 16:9 character wardrobe fitting photo for a Chinese period drama production. Left 1/3 is a face close-up showing full head to collar. Right 2/3 has three full-body standing views side by side: front, side, back."

# Age-specific imperfections
IMPERFECTIONS = {
    "15-22": "Light acne marks on forehead and chin. Visible pores on the nose. Uneven skin tone with subtle redness around the nostrils and under the eyes. Fine peach fuzz on the jawline catching side light.",
    "23-35": "Scattered freckles across the nose bridge and cheeks. Visible pores on the nose and cheeks. Uneven skin tone with subtle redness around the nostrils and under the eyes. Fine peach fuzz on the cheeks catching side light.",
    "36-50": "Crow's feet and nasolabial folds clearly visible. Visible pores across the nose and cheeks. Uneven skin tone with slight darkening under the eyes.",
    "50+": "Deep wrinkles on forehead and around eyes. Age spots near the temples and cheeks. Visible pores and sagging skin along the jawline. Uneven skin tone with darkening under the eyes.",
}

# Male stubble addition
MALE_STUBBLE = {
    "23-35": " Fine stubble grain visible on jaw.",
    "36-50": " Fine stubble grain visible on jaw.",
    "50+": " Fine facial hair on upper lip catching light.",
}

# Texture module Route A
TEXTURE_A = "Hyper-realistic raw skin texture. Bare skin, no makeup, no foundation, no retouching, no blush, no rouge on cheeks. Matte skin finish, no oil sheen, no shine, no glossy highlights on the face. {imperfections} Anti-sweet, anti-influencer aesthetic. Pure white background, facing camera, natural daylight, 58mm prime lens, f/1.4, Sony A7M4."

# Texture module Route B (no "Hyper-realistic raw skin texture" at start - it's in character line)
TEXTURE_B = "Bare skin, no makeup, no foundation, no retouching, no blush, no rouge on cheeks. Matte skin finish, no oil sheen, no shine, no glossy highlights on the face. {imperfections} Anti-sweet, anti-influencer aesthetic. Pure white background, facing camera, natural daylight, 58mm prime lens, f/1.4, Sony A7M4."

# CG trigger words check
CG_TRIGGERS = ["brocade", "gold embroidery", "gold crown", "sequin", "beaded", "iridescent", "glossy silk"]


def get_age_band(age):
    if age <= 22:
        return "15-22"
    elif age <= 35:
        return "23-35"
    elif age <= 50:
        return "36-50"
    else:
        return "50+"


def assemble_prompt(char):
    """Assemble full prompt from character variables."""
    age = char["age"]
    gender = char["gender"]  # "male" or "female"
    route = char["route"]  # "A" or "B"
    character_desc = char["character"]
    outfit_desc = char["outfit"]
    
    # Force route B if CG triggers found
    outfit_lower = outfit_desc.lower()
    if any(trigger in outfit_lower for trigger in CG_TRIGGERS):
        route = "B"
    
    age_band = get_age_band(age)
    imperfections = IMPERFECTIONS[age_band]
    
    # Add male stubble
    if gender == "male" and age_band in MALE_STUBBLE:
        imperfections += MALE_STUBBLE[age_band]
    
    if route == "A":
        # Standard three-part
        part1 = STRUCTURE_A
        part2 = f"Character: {character_desc}\n\nOutfit: {outfit_desc}"
        part3 = TEXTURE_A.format(imperfections=imperfections)
    else:
        # Route B - quality prefix in character line
        part1 = STRUCTURE_B
        # Extract key imperfection keyword for quality prefix
        if age_band == "15-22":
            quality_prefix = "with hyper-realistic raw matte skin showing visible pores and subtle acne marks"
        elif age_band == "23-35":
            quality_prefix = "with hyper-realistic raw matte skin showing visible pores and uneven tone"
        elif age_band == "36-50":
            quality_prefix = "with hyper-realistic raw matte skin showing visible pores, crow's feet and nasolabial folds"
        else:
            quality_prefix = "with hyper-realistic raw matte skin showing deep wrinkles, age spots and sagging"
        
        part2 = f"Character: {age}-year-old {gender} {quality_prefix}. {character_desc}\n\nOutfit: {outfit_desc}"
        part3 = TEXTURE_B.format(imperfections=imperfections)
    
    # For route A, character line already has age/gender in the description
    if route == "A":
        full_prompt = f"{part1}\n\n{part2}\n\n{part3}"
    else:
        full_prompt = f"{part1}\n\n{part2}\n\n{part3}"
    
    return full_prompt, route


def main():
    input_file = sys.argv[1] if len(sys.argv) > 1 else "/workspace/tmp/characters_150.json"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "/workspace/public/150_character_prompts.xlsx"
    
    chars = json.load(open(input_file, encoding="utf-8"))
    print(f"Loaded {len(chars)} characters", file=sys.stderr)
    
    # Assemble all prompts
    results = []
    for char in chars:
        prompt, final_route = assemble_prompt(char)
        results.append({**char, "prompt": prompt, "final_route": final_route})
    
    # Validate
    errors = []
    for r in results:
        # Check character description has enough features (comma count >= 7)
        comma_count = r["character"].count(",")
        if comma_count < 6:
            errors.append(f"#{r['idx']} {r['name']}: character only has {comma_count+1} features (need 8+)")
    
    if errors:
        print(f"⚠️ Validation warnings:", file=sys.stderr)
        for e in errors:
            print(f"  {e}", file=sys.stderr)
    
    # Write xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = "150 Character Prompts"
    
    # Styles
    header_fill = PatternFill("solid", fgColor="2B579A")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    wrap = Alignment(wrapText=True, vertical="top")
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )
    
    headers = ["序号", "题材", "角色名", "性别", "年龄", "身高cm", "体型", "链路", "完整Prompt"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    for i, r in enumerate(results, 2):
        ws.cell(i, 1, r["idx"])
        ws.cell(i, 2, r["genre"])
        ws.cell(i, 3, r["name"])
        ws.cell(i, 4, "男" if r["gender"] == "male" else "女")
        ws.cell(i, 5, r["age"])
        ws.cell(i, 6, r["height"])
        ws.cell(i, 7, r["body_type"])
        ws.cell(i, 8, r["final_route"])
        ws.cell(i, 9, r["prompt"])
        for c in range(1, 10):
            ws.cell(i, c).alignment = wrap
            ws.cell(i, c).border = thin_border
    
    # Column widths
    widths = [6, 12, 16, 5, 5, 7, 12, 5, 120]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+c)].width = w
    
    wb.save(output_file)
    print(f"✅ Saved: {output_file}", file=sys.stderr)
    print(f"   Total: {len(results)}", file=sys.stderr)
    print(f"   Route A: {sum(1 for r in results if r['final_route'] == 'A')}", file=sys.stderr)
    print(f"   Route B: {sum(1 for r in results if r['final_route'] == 'B')}", file=sys.stderr)
    print(f"   Male: {sum(1 for r in results if r['gender'] == 'male')}", file=sys.stderr)
    print(f"   Female: {sum(1 for r in results if r['gender'] == 'female')}", file=sys.stderr)


if __name__ == "__main__":
    main()
